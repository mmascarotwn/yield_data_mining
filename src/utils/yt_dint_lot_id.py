## test structure
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
# import pygetwindow as gw
from pyvirtualdisplay import Display

from tkinter import Tk
from tkinter.filedialog import askopenfilename

import os
import pandas as pd
from openpyxl import load_workbook


## Inputs declaration (to be changed if needed by the user)
link_web_page = 'http://mamweb.tatw.micron.com/MAMWeb/bin/MAMWeb.pl?APP=ASMTB&ACTION=REPORT&REPORTID=Status&XSLT=CLIENT&ID=368K-05:792368K&MATYPE=80&FORMAT=HTML&ORIGIN=/MAMWeb/site/ASMTB'

hbm3e_12h_id_code = "//span[text()='HBM3E (12H)']"
hbm3e_8h_id_code = "//span[text()='HBM3E (8H)']"
hbm4_12h_id_code = "//span[text()='HBM4 (12H)']"

dpm_lot_type_code = "//div[@role='option' and contains(., '6D_APTD_DPM')]"
qlc_lot_type_code = "//div[@role='option' and contains(., '6D_APTD_QLC')]"
rlc_lot_type_code = "//div[@role='option' and contains(., '6D_APTD_RLC')]"

c2pp_data_step_code = "//button[contains(text(), 'C2PP')]" # ATYMS gives all data avaialbility in raw data mode. No need to select others.

data_date_code = 'path[aria-label*="Month 202510, C2PP"]' # Change this date to be automatically detected later on


## Driver path (local PC) and windows timing setting
web_driver = r"C:\Users\mmascaro\Documents\VisualStudio Projects\venv_test_project\chromedriver.exe"
timeout = 2 # 0.5
sleep = 0

## Setup to minimise the web page on screen (WIP)
chrome_options = Options()
chrome_options.add_argument("--start-minimized") # Run in headless mode
chrome_options.add_argument("--disable-gpu") # Recommended for headless
chrome_options.add_argument("--no-sandbox") # Often required for headless mode, especially in Linux environments
chrome_options.add_argument("--disable-dev-shm-usage") # Often required for headless mode, especially in Linux environments


## Function to search for the web driver on the local PC
def chrome_web_driver(web_driver_path = None):
    """
    Text here
    """
    # Step 1: Text here
    s = Service(web_driver)
    driver = webdriver.Chrome(service = s)  # Optional argument, if not specified will search path.

    # # Minimize the Chrome window
    # for window in gw.getWindowsWithTitle('Chrome'):
    #     window.minimize()


    if not driver:
        print("Error: No driver specified/found.")
        return

    return driver

# Function to open the excel 'Open file selection' dialog to choose Excel file
def open_excel_dialogue(input = None):
    Tk().withdraw()  # Hide the root window
    excel_file = askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])

    return excel_file

# Function to read all the wafer id from the excel file 'wafer_id_in_mam' 
# & only the ones with out a target additional parameter to add yet (ex: LOT_ID)
def load_wafer_id_excel(excel_file=None):
    if not excel_file:
        print("Error: No excel_file specified/found.")
        return

    # Load the Excel sheet
    df = pd.read_excel(excel_file, sheet_name='hbm_test_yield', engine='openpyxl')

    # Check if 'LOT_ID' column exists
    lot_id_exists = 'LOT_ID' in df.columns

    # Filter wafer_id based on the presence and content of 'LOT_ID'
    if not lot_id_exists:
        wafer_id = df['wafer_id_in_mam'].dropna().astype(str).tolist()
    else:
        # Only include wafer_id where LOT_ID is empty (NaN or empty string)
        filtered_df = df[df['LOT_ID'].isna() | (df['LOT_ID'].astype(str).str.strip() == '')]
        wafer_id = filtered_df['wafer_id_in_mam'].dropna().astype(str).tolist()

    return wafer_id

## Function to generate a dynamic URL
def generate_dynamic_url(wafer_id):
    """
    Replace the static lot ID in the URL with a dynamic input.

    Parameters:
    wafer_id (str): The lot ID to insert into the URL.

    Returns:
    str: The modified URL with the new lot ID.
    """
    base_url = f"http://mamweb.tatw.micron.com/MAMWeb/bin/MAMWeb.pl?APP=ASMTB&ACTION=REPORT&REPORTID=Status&XSLT=CLIENT&ID={wafer_id}&MATYPE=80&FORMAT=HTML&ORIGIN=/MAMWeb/site/ASMTB"
    # print("base_url is:", base_url)
    
    if not wafer_id:
        print("Error: No lot_id specified/found.")
        return

    return base_url

## Function to open web page
def open_wfr_status_report_website(driver = None, link_web_page = None):
    """
    Text here
    """
    # Step 1: Text here
    driver.get(link_web_page)
    time.sleep(5) # Let the user actually see something!

    if not link_web_page :
        print("Error: No web page link found/provided.")
        return

    return 

## Function to hit enter onto the search bar (unused so far and untested)
def press_enter_in_search_bar(driver = None, timeout = None):
    """
    Locate the search bar element and simulate pressing the ENTER key.

    Parameters:
    driver (webdriver): The Selenium WebDriver instance.
    timeout (int): Maximum time to wait for the element to be present.
    """

    search_input = WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.XPATH, "//input[@type='text']")))
    search_input.send_keys(Keys.ENTER)
    # print("ENTER key pressed in the search bar.")

    return

## Function to read the Fab Lot number
def extract_fab_lot_number(driver = None, timeout = None):
    """
    Extract the FAB LOT NUMBER from the webpage.

    Parameters:
    driver (webdriver): The Selenium WebDriver instance.
    timeout (int): Maximum time to wait for the element to be present.

    Returns:
    str: The extracted FAB LOT NUMBER, or None if not found.
    """
    try:
        # Wait for the element containing 'FAB LOT NUMBER'
        fab_lot_element = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.XPATH, "//td[contains(text(), 'FAB LOT NUMBER')]/following-sibling::td"))
        )
        full_text = fab_lot_element.text.strip()
        fab_lot_number = full_text.split('.')[0]  # Get only the part before the first '.'

        print(f"FAB LOT NUMBER found: {fab_lot_number}")
        return fab_lot_number
    except Exception as e:
        print(f"Error extracting FAB LOT NUMBER: {e}")
        return None

## Function to navigate MAMASMTB(Assembly TAIWAN) and hit the search bar by entering a LOT#
def submit_lot_status_report(driver = None, timeout = None, lot_id = None):
    """
    Go to the Lot Status Report page, enter the Lot ID, and click the GO button.

    Parameters:
    driver (webdriver): Selenium WebDriver instance.
    lot_id (str): The Lot ID to submit.
    timeout (int): Timeout for waiting on elements.
    """
    try:
        # Step 1: Navigate to the Lot Status Report page
        driver.get("http://mamweb.tatw.micron.com/MAMWeb/site/ASMTB/")

        # Step 2: Wait for the input box and enter the Lot ID
        input_box = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.NAME, "ID"))
        )
        input_box.clear()
        input_box.send_keys(lot_id)

        # Step 3: Click the GO button
        go_button = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.NAME, "btnGo"))
        )
        go_button.click()

        print(f"Submitted Lot ID '{lot_id}' successfully.")
    except Exception as e:
        print(f"Error submitting Lot Status Report: {e}")
    
    return

## Function to match the wafers and find the right LOT_ID on MAMASMTB
## TO-DO: 1) separate the excel save function from the matching; 2) add time stamp to the file.
def find_matching_lot_and_update_excel(driver=None, timeout=None, wafer_id_in_mam=None, excel_file_path=None):
    """
    Searches through LOT IDs and updates the Excel file with the matching LOT ID for the given wafer ID.
    """
    try:
        # Step 1: Extract all LOT IDs from the dropdown
        # select_element = WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.NAME, "ID")))
        select_element = WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.XPATH, "//select[@name='ID']")))
        options = select_element.find_elements(By.TAG_NAME, "option")

        lot_ids = [opt.get_attribute("value") for opt in options if opt.get_attribute("value")]
        print("Available LOT IDs:", lot_ids) 

        # Step 2: Iterate through each LOT ID
        for lot_id in lot_ids:
            Select(driver.find_element(By.NAME, "ID")).select_by_value(lot_id)
            print(f"Selected LOT ID: {lot_id}")

            # Step 3: Click the SUBMIT button 
            try:
                submit_button = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((By.XPATH, "//input[@type='SUBMIT']")))
                driver.execute_script("arguments[0].scrollIntoView(true);", submit_button)
                WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((By.NAME, "SUBMIT")))
                submit_button.click()
                # print("Clicked SUBMIT button.")
                time.sleep(0.5)

            except Exception as e:
                # print(f"Standard click failed: {e}. Trying JS click.")
                try:
                    driver.execute_script("arguments[0].click();", submit_button)
                    # print("Clicked SUBMIT button via JS.")
                except Exception as js_e:
                    # print(f"JavaScript click also failed: {js_e}")
                    continue  # Skip to next LOT ID

            # Step 4: Extract wafer numbers from the new page
            wafer_elements = WebDriverWait(driver, timeout).until(EC.presence_of_all_elements_located((By.XPATH, "//tr[@class='ReportRow1' or @class='ReportRow0']//td[@class='ReportCellAlt']")))
            wafer_numbers = [el.text.strip() for el in wafer_elements]
            print("Extracted wafer numbers:", wafer_numbers)

            # Step 5: Check for match
            if wafer_id_in_mam in wafer_numbers:
                print(f"Match found: {lot_id}")

            # Step 6: Update Excel file
                df = pd.read_excel(excel_file_path, sheet_name='hbm_test_yield', engine='openpyxl')
                if 'LOT_ID' not in df.columns:
                    df['LOT_ID'] = None

                # Ensure matching works even if types differ
                df['wafer_id_in_mam'] = df['wafer_id_in_mam'].astype(str)
                wafer_id_in_mam = str(wafer_id_in_mam)

                # Update the matching row
                df.loc[df['wafer_id_in_mam'] == wafer_id_in_mam, 'LOT_ID'] = lot_id

                # Save back to the same sheet without affecting other sheets
                from openpyxl import load_workbook
                book = load_workbook(excel_file_path)

                with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    writer._book = book
                    writer._sheets = {ws.title: ws for ws in book.worksheets}
                    df.to_excel(writer, sheet_name='hbm_test_yield', index=False)

            # # Step 7: Rename the file if needed
            #     base, ext = os.path.splitext(excel_file_path)
            #     if not base.endswith('integrated_data'):
            #         new_file_path = f"{base}_integrated_data{ext}"
            #         os.rename(excel_file_path, new_file_path)
            #         print(f"File renamed to: {new_file_path}")
            #     else:
            #         print("Filename already includes 'integrated_data'. No renaming needed.")

                break

            # Step 8: Go back to the LOT selection page
            driver.back()

    except Exception as e:
        print(f"Error in matching LOT ID and updating Excel: {e}")
    
    return

## Function to quit the process afterwards
def quit_script(driver = None, timeout = None):
    """
    Wait and Quit script
    """
    time.sleep(10) # Let the user actually see something!
    driver.quit()

    if not timeout:
        print("Error: No timeout specified.")
        return

    return 


###
### Functions grouping
###

def lot_search(a=None):
    select_excel_file = open_excel_dialogue()
    read_wafer_id = load_wafer_id_excel(select_excel_file)
    print(read_wafer_id)

    for wafer in read_wafer_id:
        print(f"Processing wafer ID: {wafer}")

        generale_url = generate_dynamic_url(wafer)  # Pass single wafer ID
        driver = chrome_web_driver(web_driver)
        dynamic_link_web_page = generate_dynamic_url(generale_url)
        wafer_MAM_web_page = open_wfr_status_report_website(driver, dynamic_link_web_page)
        read_fab_lot_number = extract_fab_lot_number(driver, timeout)
        search_lot_MAM = submit_lot_status_report(driver, timeout, read_fab_lot_number)
        search_wafers_match_MAM = find_matching_lot_and_update_excel(driver, timeout, wafer, select_excel_file)
        quit = quit_script(driver, timeout)

    # # Rename the Excel file only once after all operations
    # base, ext = os.path.splitext(select_excel_file)
    # if not base.endswith('integrated_data'):
    #     new_file_path = f"{base}_integrated_data{ext}"
    #     os.rename(select_excel_file, new_file_path)
    #     print(f"Excel file renamed to: {new_file_path}")
    # else:
    #     print("Excel file already includes 'integrated_data'. No renaming needed.")


###
### Test the functions
###
lot_search()